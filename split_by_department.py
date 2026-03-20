#!/usr/bin/env python3
"""
Tách file tổng thành 42 file theo phòng ban (cột AJ - col index 35).
Giữ nguyên template gốc (rows 0-38), chỉ copy data rows tương ứng.
"""
import datetime
import os

import xlrd
from xlrd import xldate_as_tuple
import xlwt
from xlutils.copy import copy as xl_copy
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

# Lấy file đầu tiên từ thư mục input
input_dir = os.path.join(os.path.dirname(__file__), 'input')
allowed_exts = {'.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'}
files = [f for f in os.listdir(input_dir)
         if os.path.isfile(os.path.join(input_dir, f))
         and os.path.splitext(f)[1].lower() in allowed_exts]
# print(f"Tìm thấy {len(files)} file trong thư mục input.")
# exit()

if not files:
    raise ValueError("Không có file nào trong thư mục input")
files.sort()  # Sắp xếp theo thứ tự alphabet
SRC = os.path.join(input_dir, files[0])

# print(SRC)
# exit()

OUT_DIR = os.path.join(os.path.dirname(__file__), 'output')
os.makedirs(OUT_DIR, exist_ok=True)

def _load_workbook_any(path):
    """Load a workbook from .xlsx/.xls, returning an openpyxl Workbook.

    For .xls input we convert it to openpyxl Workbook (values only),
    because openpyxl does not support .xls directly.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        rb = xlrd.open_workbook(path, formatting_info=True)
        wb = Workbook()
        # Remove default sheet created by openpyxl
        default = wb.active
        wb.remove(default)

        for sheet_name in rb.sheet_names():
            sh = rb.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)

            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = datetime.datetime(*xldate_as_tuple(cell.value, rb.datemode)[:6])
                            ws.cell(row=r + 1, column=c + 1, value=dt)
                        except Exception:
                            ws.cell(row=r + 1, column=c + 1, value=cell.value)
                    else:
                        ws.cell(row=r + 1, column=c + 1, value=cell.value)

        return wb

    return load_workbook(path, data_only=False)


def _is_xls_date(cell, rb):
    """Return True if xlrd cell should be treated as a date (by format or type)."""
    if cell.ctype == xlrd.XL_CELL_DATE:
        return True
    if cell.ctype != xlrd.XL_CELL_NUMBER:
        return False

    xf = rb.xf_list[cell.xf_index]
    fmt = rb.format_map.get(xf.format_key, "")
    if hasattr(fmt, 'format_str'):
        fmt_str = fmt.format_str
    else:
        fmt_str = str(fmt)
    fmt_lower = fmt_str.lower()
    # crude check: date formatting contains y/m/d or h/s
    return any(tok in fmt_lower for tok in ('yy', 'mm', 'dd', 'hh', 'ss'))


def _parse_text_date(s):
    """Try to parse common date strings (dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd)."""
    if not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d.%m.%Y'):
        try:
            return datetime.datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def _process_xls(SRC, dept_rows, input_ext, out_ext, src_sheet, src_max_col, src_max_row):
    """Process .xls input by using xlutils to preserve formatting and all sheets."""
    # xlutils.copy (xlrd -> xlwt) keeps formatting, merged cells, etc.
    rb = xlrd.open_workbook(SRC, formatting_info=True)
    sheet_names = rb.sheet_names()
    main_index = sheet_names.index(src_sheet.title)

    for dept_name, rows in sorted(dept_rows.items()):
        wb_new = xl_copy(rb)
        ws = wb_new.get_sheet(main_index)

        # Clear existing data rows (keep header/template rows intact)
        for r in range(DATA_START, src_max_row):
            for c in range(src_max_col):
                ws.write(r, c, '')

        # Write data rows starting after header
        write_row = DATA_START
        for src_r in rows:
            for c in range(src_max_col):
                cell = rb.sheet_by_index(main_index).cell(src_r - 1, c)
                val = cell.value

                # If the cell is text but looks like a date string, parse & write as date
                parsed = _parse_text_date(val)
                if parsed:
                    ws.write(write_row, c, parsed.strftime('%d/%m/%Y'))
                    continue

                if _is_xls_date(cell, rb):
                    try:
                        dt = datetime.datetime(*xldate_as_tuple(val, rb.datemode)[:6])
                        ws.write(write_row, c, dt.strftime('%d/%m/%Y'))
                    except Exception:
                        ws.write(write_row, c, val)
                else:
                    ws.write(write_row, c, val)
            write_row += 1

        fname = safe_filename(dept_name) + out_ext
        out_path = os.path.join(OUT_DIR, fname)
        wb_new.save(out_path)
        
        if IS_PDF:
            pdf_fname = safe_filename(dept_name) + '.pdf'
            pdf_path = os.path.join(OUT_DIR, pdf_fname)
            if _convert_excel_to_pdf(out_path, pdf_path):
                os.remove(out_path)  # Remove Excel file after successful conversion
                print(f"  ✓ {pdf_fname} ({len(rows)} dòng)")
            else:
                print(f"  ✓ {fname} ({len(rows)} dòng)")
        else:
            print(f"  ✓ {fname} ({len(rows)} dòng)")


# --- Read source ---
wb_src = _load_workbook_any(SRC)
SHEET_NAME = wb_src.sheetnames[0]  # Lấy sheet đầu tiên
src_sheet = wb_src[SHEET_NAME]

# DEPT_COL = 35        # AJ (0-based)
# HEADER_END = 38      # rows 0-38 are template/header (0-based)
# DATA_START = 39      # data begins at row 39 (0-based)
# IS_PDF = False        # If True, output PDF files instead of Excel (requires libreoffice)

DEPT_COL = 35        # AJ (0-based)
HEADER_END = 7      # rows 0-38 are template/header (0-based)
DATA_START = 8      # data begins at row 39 (0-based)
IS_PDF = False

# DEPT_COL = 26        # AA (0-based)
# HEADER_END = 5      # rows 0-38 are template/header (0-based)
# DATA_START = 6      # data begins at row 39 (0-based)
# IS_PDF = False

# OpenPyXL is 1-based for rows/cols
DEPT_COL_IDX = DEPT_COL + 1
DATA_START_ROW = DATA_START + 1
HEADER_END_ROW = HEADER_END + 1

# Collect data rows grouped by department
dept_rows = {}  # {dept_name: [row_indices (1-based)]}
for r in range(DATA_START_ROW, src_sheet.max_row + 1):
    dept = src_sheet.cell(row=r, column=DEPT_COL_IDX).value
    if not dept:
        dept = '_Không có phòng ban'
    dept = str(dept).strip()
    dept_rows.setdefault(dept, []).append(r)

print(f"Tổng phòng ban: {len(dept_rows)}")
print(f"Tổng dòng data: {sum(len(v) for v in dept_rows.values())}")

# exit()

def safe_filename(name):
    """Tạo tên file an toàn từ tên phòng ban."""
    return name.replace('/', '-').replace('\\', '-').replace(':', '-')


def _get_vietnamese_font():
    """Find and register a TrueType font that supports Vietnamese characters."""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # Common font paths on macOS
        font_paths = [
            '/System/Library/Fonts/Arial.ttf',
            '/System/Library/Fonts/Helvetica.ttc',
            '/Library/Fonts/Arial.ttf',
            '/Library/Fonts/DejaVuSans.ttf',
            '/System/Library/Fonts/Tahoma.ttf',
            '/System/Library/Fonts/Georgia.ttf',
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('VietFont', font_path))
                    return 'VietFont'
                except Exception:
                    continue
        
        return 'Courier'  # Fallback to Courier if no suitable font found
    except Exception:
        return 'Courier'


def _convert_excel_to_pdf(excel_path, pdf_path):
    """Convert Excel file to PDF using Python libraries. Returns True if successful."""
    try:
        from reportlab.lib import pagesizes
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
        from reportlab.lib.units import inch, cm
        from reportlab.lib import colors
        from openpyxl.utils import get_column_letter
        
        # Get Vietnamese-compatible font
        font_name = _get_vietnamese_font()
        
        # Extract data from Excel based on file format
        data = []
        hidden_cols = set()  # Track which column indices are hidden
        col_widths_original = {}  # Store original column widths from Excel
        ext = os.path.splitext(excel_path)[1].lower()
        
        if ext == '.xls':
            # Use xlrd for .xls files
            rb = xlrd.open_workbook(excel_path, formatting_info=True)
            sheet = rb.sheet_by_index(0)
            
            # Check for hidden columns and get column widths using sheet's colinfo_map
            for col_idx in range(sheet.ncols):
                colinfo = sheet.colinfo_map.get(col_idx) if hasattr(sheet, 'colinfo_map') else None
                if colinfo:
                    if colinfo.hidden:
                        hidden_cols.add(col_idx)
                    # Store width in twips (1/20 of a point), convert to cm
                    col_widths_original[col_idx] = (colinfo.width / 20.0) * 0.0352778 if colinfo.width else 2.8
                else:
                    # Default width for Excel is approximately 8.43 characters = 2.8 cm
                    col_widths_original[col_idx] = 2.8
            
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    if col_idx not in hidden_cols:  # Skip hidden columns
                        cell = sheet.cell(row_idx, col_idx)
                        row_data.append(str(cell.value) if cell.value is not None else '')
                data.append(row_data)
        else:
            # Use openpyxl for .xlsx and other formats
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            # Get list of hidden columns and column widths from column_dimensions
            col_display_idx = 0  # Track visible column index
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                col_dim = ws.column_dimensions.get(col_letter)
                
                if col_dim and col_dim.hidden:
                    hidden_cols.add(col_idx - 1)
                
                # Width in openpyxl is in character units, convert to cm (1 char ≈ 0.21 cm)
                if col_dim and col_dim.width:
                    col_widths_original[col_idx - 1] = col_dim.width * 0.21
                else:
                    col_widths_original[col_idx - 1] = 2.55  # Default width
            
            for row in ws.iter_rows(values_only=True):
                row_data = []
                for col_idx, cell in enumerate(row):
                    if col_idx not in hidden_cols:  # Skip hidden columns
                        row_data.append(str(cell) if cell is not None else '')
                data.append(row_data)
        
        if not data:
            print(f"  ⚠ File Excel trống: {excel_path}")
            return False
        
        # Create PDF with landscape orientation and smaller margins
        pagesize = pagesizes.landscape(pagesizes.A4)
        doc = SimpleDocTemplate(pdf_path, pagesize=pagesize, leftMargin=0.3*cm, 
                               rightMargin=0.3*cm, topMargin=0.5*cm, bottomMargin=0.5*cm)
        elements = []
        
        # Define table style
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ])
        
        # Use all visible columns
        table_data = data
        
        # Use original column widths from Excel, scaled proportionally
        page_width = pagesize[0] - 0.6*cm  # Account for margins
        
        # Calculate visible column widths
        col_widths = []
        visible_col_idx = 0
        for col_idx in range(sheet.ncols if ext == '.xls' else ws.max_column):
            if col_idx not in hidden_cols:
                width = col_widths_original.get(col_idx, 2.8)
                col_widths.append(width * cm)
        
        # If total width exceeds page width, scale down proportionally
        total_width = sum(col_widths)
        if total_width > page_width:
            scale_factor = page_width / total_width
            col_widths = [w * scale_factor for w in col_widths]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(style)
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        return True
        
    except ImportError as e:
        print(f"  ⚠ Thiếu thư viện Python: {e}")
        print(f"    Cài đặt bằng: pip install reportlab")
        return False
    except Exception as e:
        print(f"  ⚠ Lỗi chuyển PDF: {e}")
        return False


def _copy_sheet(src, dst):
    """Copy values/styles/merged-cells/column widths from src sheet to dst sheet."""
    # Column widths
    for col_letter, dim in src.column_dimensions.items():
        dst.column_dimensions[col_letter].width = dim.width
        dst.column_dimensions[col_letter].hidden = dim.hidden

    # Row heights
    for row, dim in src.row_dimensions.items():
        dst.row_dimensions[row].height = dim.height

    # Cell values (styles dropped to avoid lost/invalid style indices on save)
    for row in src.iter_rows():
        for cell in row:
            # Skip placeholder cells that are part of a merged range.
            if isinstance(cell, MergedCell):
                continue

            new_cell = dst.cell(row=cell.row, column=cell.col_idx, value=cell.value)
            if cell.number_format:
                new_cell.number_format = cell.number_format
            if cell.hyperlink:
                new_cell.hyperlink = cell.hyperlink
            if cell.comment:
                new_cell.comment = cell.comment

    # Merged cells
    for merged in src.merged_cells.ranges:
        try:
            dst.merge_cells(str(merged))
        except Exception:
            # In some cases OpenPyXL may error when copying complex merged-cell styles.
            # We can ignore these merge ranges to keep the script working.
            pass


# --- For each department, create a new workbook (keep all sheets + formatting) ---
input_ext = os.path.splitext(SRC)[1].lower()
# If input was .xls, output as .xls; otherwise keep same extension
out_ext = input_ext if input_ext != '.xls' else '.xls'

src_max_col = src_sheet.max_column
src_max_row = src_sheet.max_row

if input_ext == '.xls':
    # For .xls, use xlutils (xlrd -> xlwt) to preserve formatting.
    _process_xls(SRC, dept_rows, input_ext, out_ext, src_sheet, src_max_col, src_max_row)
else:
    for dept_name, rows in sorted(dept_rows.items()):
        # Reload original workbook to preserve all sheets/styles unchanged
        wb_new = _load_workbook_any(SRC)
        ws = wb_new[SHEET_NAME]

        # Clear existing data rows (keep header/template rows intact)
        for r in range(DATA_START_ROW, src_max_row + 1):
            for c in range(1, src_max_col + 1):
                ws.cell(row=r, column=c).value = None

        # Write data rows starting after header
        write_row = DATA_START_ROW
        for src_r in rows:
            for c in range(1, src_max_col + 1):
                src_cell = src_sheet.cell(row=src_r, column=c)
                tgt_cell = ws.cell(row=write_row, column=c)
                tgt_cell.value = src_cell.value
                if src_cell.number_format:
                    tgt_cell.number_format = src_cell.number_format
                if src_cell.hyperlink:
                    tgt_cell.hyperlink = src_cell.hyperlink
                if src_cell.comment:
                    tgt_cell.comment = src_cell.comment
            write_row += 1

        fname = safe_filename(dept_name) + out_ext
        out_path = os.path.join(OUT_DIR, fname)
        wb_new.save(out_path)
        
        if IS_PDF:
            pdf_fname = safe_filename(dept_name) + '.pdf'
            pdf_path = os.path.join(OUT_DIR, pdf_fname)
            if _convert_excel_to_pdf(out_path, pdf_path):
                os.remove(out_path)  # Remove Excel file after successful conversion
                print(f"  ✓ {pdf_fname} ({len(rows)} dòng)")
            else:
                print(f"  ✓ {fname} ({len(rows)} dòng)")
        else:
            print(f"  ✓ {fname} ({len(rows)} dòng)")

print(f"\nHoàn tất! {len(dept_rows)} file đã lưu vào: {OUT_DIR}")
